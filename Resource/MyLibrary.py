class Data:

    def get_data_by_sheet_name(self, sFilePath, sSheetName, encode):
        # """Get data from given file path, sheet name then return array of object"""
        from openpyxl import load_workbook
        # noinspection PyUnresolvedReferences
        from openpyxl import Workbook

        wb = load_workbook(sFilePath.strip())
        ws = wb.get_sheet_by_name(sSheetName)
        # ws = wb.worksheets(sSheetName)
        row_count = ws.max_row
        column_count = ws.max_column

        # init data returned
        test_data = []

        for i in range(2, row_count + 1):
            # ini dict
            row_data = {}
            for j in range(1, column_count + 1):
                if str(ws.cell(row=1, column=j).value) == 'None':
                    continue
                else:
                    key = ws.cell(row=1, column=j).value
                    value = ws.cell(row=i, column=j).value
                    if value == None:
                        value = ''
                    if encode == 'False':
                        row_data[key] = str(value)
                    if encode == 'True':
                        row_data[key] = str(value.encode('utf-8'))
            # print row_data
            test_data.append(row_data.copy())

        return test_data
    # end get_data_by_sheet_name

    def change_email(self, email):
        x1 = email.split("@")
        x2 = x1[0].split("+")
        x3 = str(int(x2[1]) + 1)
        new_email = email.replace(x2[1], x3)
        return new_email
